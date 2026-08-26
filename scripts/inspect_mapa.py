#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Inspeciona aba MAPA DE VIATURAS (possivel lista mestra)"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

XLSX = r"D:\USER\DESKTOPP\excel\MAPA GERAL CPI-7 - 2026.xlsx"
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["MAPA DE VIATURAS"]

print(f"ABA MAPA DE VIATURAS ({ws.max_row} linhas x {ws.max_column} cols)")
print("=" * 80)
# Header
for row in range(1, 10):
    cells = []
    for col in range(1, min(15, ws.max_column + 1)):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            vstr = str(v)[:30]
            cells.append(f"[{col}]={vstr}")
    if cells:
        print(f"L{row:3d}: " + " | ".join(cells))
print()
print("...")
print("=" * 80)
# Linhas com dados (vou achar)
contador = 0
for row in range(11, ws.max_row + 1):
    seq = ws.cell(row=row, column=1).value
    if seq is None or not isinstance(seq, (int, float)):
        continue
    contador += 1
    if contador <= 20 or contador > 800:  # mostra primeiras 20 e depois de 800
        if contador > 20 and contador <= 800: continue
        cells = []
        for col in range(1, min(15, ws.max_column + 1)):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                vstr = str(v)[:25]
                cells.append(f"[{col}]={vstr}")
        print(f"L{row:3d}: " + " | ".join(cells))

print()
print(f"Total de linhas com SEQ numerico: {contador}")
