#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Importa viaturas em paralelo (10 workers)"""
import sys, io, json, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import urllib.request
import urllib.error
import openpyxl
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

XLSX = r"D:\USER\DESKTOPP\excel\MAPA GERAL CPI-7 - 2026.xlsx"
CONVEX_URL = "http://10.36.177.138:3212"
WILLIAM_CPF = "26034202833"
WORKERS = 10

UNITS = {
    "CPI7":   {"code": "607000000", "id": "j973x3z4qna5vt761edcgj2rjn8c0zzy", "sigla": "CPI-7"},
    "7BPMI":  {"code": "607070000", "id": "j9784whpmfq1zf3fgehr1y0s6n8c1rfz", "sigla": "7BPMI"},
    "12BPMI": {"code": "607120000", "id": "j973w824vf35z5rcsvjecjz8vx8c1qyw", "sigla": "12BPMI"},
    "22BPMI": {"code": "607220000", "id": "j97dt9rpszb3m3gzq93dq8rgwh8c1262", "sigla": "22BPMI"},
    "40BPMI": {"code": "607400000", "id": "j973qv18706btfrynsmfv3a4r98c0gyy", "sigla": "40BPMI"},
    "50BPMI": {"code": "607500000", "id": "j979rgpe5jnwj2sgv6rzw22gn18c1zx8", "sigla": "50BPMI"},
    "53BPMI": {"code": "607530000", "id": "j97eh5eqbezbb8cgq86jvkeacn8c0cr8", "sigla": "53BPMI"},
    "54BPMI": {"code": "607540000", "id": "j971pm50zmw8a538fb4fcwg0n58c116k", "sigla": "54BPMI"},
    "55BPMI": {"code": "607550000", "id": "j97035xc55veqydfmk7gkzz5px8c0zbd", "sigla": "55BPMI"},
    "14BAEP": {"code": "607140000", "id": "j977mhwvmhpsr69nqt07p8x35d8c1w8y", "sigla": "14BAEP"},
}


def call_mutation(path, args):
    url = f"{CONVEX_URL}/api/mutation"
    body = json.dumps({"path": path, "args": args}).encode()
    req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        body = e.read().decode(errors='replace')[:300]
        return {"status": "error", "errorMessage": f"HTTP {e.code}: {body}"}
    except Exception as e:
        return {"status": "error", "errorMessage": str(e)}


def parse_data_baixa(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return int(value.timestamp() * 1000)
    if isinstance(value, (int, float)):
        return int((value - 25569) * 86400 * 1000)
    if isinstance(value, str):
        try:
            d = datetime.strptime(value.strip(), "%d/%m/%Y")
            return int(d.timestamp() * 1000)
        except:
            return None
    return None


def collect_viaturas(wb):
    """Le xlsx e retorna lista de (unidade, dados)"""
    items = []
    for unit_name, unit_info in UNITS.items():
        if unit_name not in wb.sheetnames:
            continue
        ws = wb[unit_name]
        for row in range(2, ws.max_row + 1):
            seq = ws.cell(row=row, column=1).value
            tipo = ws.cell(row=row, column=2).value
            categoria = ws.cell(row=row, column=3).value
            data_baixa_raw = ws.cell(row=row, column=4).value
            prefixo = ws.cell(row=row, column=5).value
            marca_modelo = ws.cell(row=row, column=6).value
            motivo = ws.cell(row=row, column=8).value
            situação = ws.cell(row=row, column=12).value
            if not prefixo or not isinstance(prefixo, str):
                continue
            if not tipo or tipo not in ("MT", "CR"):
                continue
            cat = (categoria or "").strip().upper()
            if cat not in ("OPERACIONAL", "ADM"):
                cat = "OPERACIONAL"
            data_baixa = parse_data_baixa(data_baixa_raw)
            # FIX (William 2026-08-07): a planilha é "VTR BAIXADAS" - TODAS as viaturas
            # das abas de unidade estao em alguma situação de baixa (manutencao, patio,
            # aguardando pregao, etc). A coluna CATEGORIA (OPERACIONAL/ADM) NAO
            # indica se ta ativa, e sim o TIPO DE USO.
            # Regra: se a viatura ta na planilha = baixada. Só fica ativa se
            # EXPLICITAMENTE tiver dataBaixa vazia E nenhum motivo (improvavel).
            if motivo or (situacao and isinstance(situacao, str) and any(
                kw in situacao.upper() for kw in
                ["PATIO", "MANUTENCAO", "MANUTENÇÃO", "AGUARDANDO", "SINDICANCIA", "SINDICÂNCIA", "SERVICO", "SERVIÇO", "BTL"]
            )):
                ativo = False
            elif data_baixa is None and not motivo and not situacao:
                # Linha vazia (sem motivo, sem situacao, sem data) - provavelmente placeholder
                continue
            else:
                ativo = False  # default: planilha só tem baixadas
            args = {
                "cpf": WILLIAM_CPF,
                "opm": unit_info["id"],
                "prefixo": prefixo.strip(),
                "tipo": tipo,
                "categoria": cat,
                "marcaModelo": str(marca_modelo or "").strip(),
                "ativo": ativo,
            }
            if data_baixa: args["dataBaixa"] = data_baixa
            if motivo: args["motivo"] = str(motivo).strip()
            if situacao: args["situacao"] = str(situacao).strip()
            items.append((unit_info["sigla"], prefixo.strip(), args))
    return items


def process_one(item):
    sigla, prefixo, args = item
    res = call_mutation("viaturas:upsert", args)
    if res.get("status") == "success":
        created = res.get("value", {}).get("created", False)
        return ("ok", sigla, prefixo, created)
    return ("err", sigla, prefixo, res.get("errorMessage", "")[:200])


def main():
    print("Lendo xlsx...")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    items = collect_viaturas(wb)
    print(f"Total de viaturas a importar: {len(items)}")
    print()

    t0 = time.time()
    criadas = atualizadas = erros = 0
    por_unidade = {}
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futures = [ex.submit(process_one, item) for item in items]
        done = 0
        for fut in as_completed(futures):
            done += 1
            status, sigla, prefixo, info = fut.result()
            por_unidade.setdefault(sigla, {"criadas": 0, "atualizadas": 0, "erros": 0})
            if status == "ok":
                if info:
                    criadas += 1
                    por_unidade[sigla]["criadas"] += 1
                else:
                    atualizadas += 1
                    por_unidade[sigla]["atualizadas"] += 1
            else:
                erros += 1
                por_unidade[sigla]["erros"] += 1
                if erros <= 5:
                    print(f"  ERRO {sigla}/{prefixo}: {info}")
            if done % 25 == 0:
                elapsed = time.time() - t0
                print(f"  ...{done}/{len(items)} ({elapsed:.1f}s)")

    elapsed = time.time() - t0
    print()
    print("=" * 50)
    for sigla in sorted(por_unidade.keys()):
        d = por_unidade[sigla]
        print(f"  {sigla:<8} criadas={d['criadas']:3d} atualizadas={d['atualizadas']:3d} erros={d['erros']}")
    print(f"\nTOTAL: criadas={criadas} atualizadas={atualizadas} erros={erros} em {elapsed:.1f}s")


if __name__ == "__main__":
    main()
