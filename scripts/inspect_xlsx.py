#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Inspeciona estrutura do xlsx MAPA GERAL CPI-7"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'], check=True)
    import openpyxl

XLSX = r"D:\USER\DESKTOPP\excel\MAPA GERAL CPI-7 - 2026.xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)
print(f"PLANILHA: {XLSX}")
print(f"TOTAL DE ABAS: {len(wb.sheetnames)}")
print()
for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    rows = ws.max_row
    cols = ws.max_column
    print(f"  [{i+1}] '{name}': {rows} linhas x {cols} cols")

print()
print("=" * 60)
print("DETALHES DA PRIMEIRA ABA (capa)")
print("=" * 60)
ws = wb[wb.sheetnames[0]]
for row in range(1, min(20, ws.max_row + 1)):
    cells = []
    for col in range(1, min(8, ws.max_column + 1)):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            cells.append(f"[{row},{col}]={v!r}")
    if cells:
        print(" | ".join(cells))
