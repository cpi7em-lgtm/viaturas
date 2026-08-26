#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Inspeciona aba de uma unidade"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

XLSX = r"D:\USER\DESKTOPP\excel\MAPA GERAL CPI-7 - 2026.xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["7BPMI"]

print(f"ABA: 7BPMI ({ws.max_row} linhas x {ws.max_column} cols)")
print("=" * 80)
# Mostra primeiras 30 linhas pra ver header + primeiras viaturas
for row in range(1, 30):
    cells = []
    for col in range(1, min(15, ws.max_column + 1)):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            vstr = str(v)[:30]
            cells.append(f"[{col}]={vstr}")
    if cells:
        print(f"L{row:3d}: " + " | ".join(cells))
print()
print("=" * 80)
# Mostra ultimas 10 linhas (geralmente tem TOTAL)
for row in range(max(1, ws.max_row - 10), ws.max_row + 1):
    cells = []
    for col in range(1, min(15, ws.max_column + 1)):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            vstr = str(v)[:30]
            cells.append(f"[{col}]={vstr}")
    if cells:
        print(f"L{row:3d}: " + " | ".join(cells))
