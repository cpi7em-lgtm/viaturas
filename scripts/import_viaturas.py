#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Importa viaturas do MAPA GERAL CPI-7.xlsx pro Convex"""
import sys, io, json, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import urllib.request
import urllib.error
import openpyxl
from datetime import datetime

XLSX = r"D:\USER\DESKTOPP\excel\MAPA GERAL CPI-7 - 2026.xlsx"
CONVEX_URL = "http://localhost:3212"  # API direta
WILLIAM_CPF = "26034202833"

# Mapping: nome da aba -> {code SIAFEM, id Convex}
UNITS = {
    "CPI7":   {"code": "607000000", "id": "j976xav14pysg2bvqbmekqrr3s8c025a", "sigla": "CPI-7"},
    "7BPMI":  {"code": "607070000", "id": "j973zyhya6h0f3dx73fm3t470d8c0x81", "sigla": "7BPMI"},
    "12BPMI": {"code": "607120000", "id": "j976pbt7eqa9nz59q0gjhfnnvn8c1ah5", "sigla": "12BPMI"},
    "22BPMI": {"code": "607220000", "id": "j978qb27e2hg2a2ff66krp4r3x8c0jne", "sigla": "22BPMI"},
    "40BPMI": {"code": "607400000", "id": "j97efmqa4j1tmfewwrpy6fc44h8c1p1a", "sigla": "40BPMI"},
    "50BPMI": {"code": "607500000", "id": "j9786bsk65s8441thf4mj5m1398c14qf", "sigla": "50BPMI"},
    "53BPMI": {"code": "607530000", "id": "j974bveht474kmt6jb9tx9kbwd8c0ewb", "sigla": "53BPMI"},
    "54BPMI": {"code": "607540000", "id": "j977wbrw7w6d3gzvjy7fhc69n58c1mh3", "sigla": "54BPMI"},
    "55BPMI": {"code": "607550000", "id": "j972r6v1vyfexw8zmqsb2dbn3h8c0888", "sigla": "55BPMI"},
    "14BAEP": {"code": "607140000", "id": "j975t4y8qsqk32a9dd7s20rz9x8c057e", "sigla": "14BAEP"},
}


def call_mutation(path, args):
    url = f"{CONVEX_URL}/api/mutation"
    body = json.dumps({"path": path, "args": args}).encode()
    req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        body = e.read().decode(errors='replace')[:500]
        return {"status": "error", "errorMessage": f"HTTP {e.code}: {body}"}
    except Exception as e:
        return {"status": "error", "errorMessage": str(e)}


def parse_data_baixa(value):
    """Converte data do Excel pra timestamp em ms. None se vazio."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return int(value.timestamp() * 1000)
    if isinstance(value, (int, float)):
        # Excel data: número de dias desde 1900-01-01
        return int((value - 25569) * 86400 * 1000)
    if isinstance(value, str):
        try:
            # tenta BR format: dd/mm/yyyy
            d = datetime.strptime(value.strip(), "%d/%m/%Y")
            return int(d.timestamp() * 1000)
        except:
            return None
    return None


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    total_imported = 0
    total_updated = 0
    total_errors = 0
    por_unidade = {}

    for unit_name, unit_info in UNITS.items():
        if unit_name not in wb.sheetnames:
            print(f"AVISO: aba '{unit_name}' não encontrada")
            continue
        ws = wb[unit_name]
        criadas = atualizadas = erros = 0
        primeira_linha = 2  # linha 1 é header

        for row in range(primeira_linha, ws.max_row + 1):
            seq = ws.cell(row=row, column=1).value
            tipo = ws.cell(row=row, column=2).value
            categoria = ws.cell(row=row, column=3).value
            data_baixa_raw = ws.cell(row=row, column=4).value
            prefixo = ws.cell(row=row, column=5).value
            marca_modelo = ws.cell(row=row, column=6).value
            motivo = ws.cell(row=row, column=8).value
            situação = ws.cell(row=row, column=12).value

            if not prefixo or not isinstance(prefixo, str):
                continue
            if not tipo or tipo not in ("MT", "CR"):
                continue
            cat = (categoria or "").strip().upper()
            if cat not in ("OPERACIONAL", "ADM"):
                cat = "OPERACIONAL"  # default

            data_baixa = parse_data_baixa(data_baixa_raw)
            ativo = data_baixa is None  # sem data = operando

            args = {
                "cpf": WILLIAM_CPF,
                "opm": unit_info["id"],
                "prefixo": prefixo.strip(),
                "tipo": tipo,
                "categoria": cat,
                "marcaModelo": str(marca_modelo or "").strip(),
                "ativo": ativo,
            }
            if data_baixa:
                args["dataBaixa"] = data_baixa
            if motivo:
                args["motivo"] = str(motivo).strip()
            if situacao:
                args["situacao"] = str(situacao).strip()

            res = call_mutation("viaturas:upsert", args)
            if res.get("status") == "success":
                if res.get("value", {}).get("created"):
                    criadas += 1
                    total_imported += 1
                else:
                    atualizadas += 1
                    total_updated += 1
            else:
                erros += 1
                total_errors += 1
                if erros <= 3:  # só mostra primeiros 3 erros
                    print(f"  ERRO {unit_name} {prefixo}: {res.get('errorMessage', res)[:200]}")

        por_unidade[unit_info["sigla"]] = {"criadas": criadas, "atualizadas": atualizadas, "erros": erros}
        print(f"  {unit_info['sigla']:<8} criadas={criadas:3d} atualizadas={atualizadas:3d} erros={erros}")

    print()
    print("=" * 50)
    print(f"TOTAL: importadas={total_imported} atualizadas={total_updated} erros={total_errors}")
    print(f"Unidades: {por_unidade}")


if __name__ == "__main__":
    main()
