#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Conta viaturas reais e ve as ultimas linhas uteis"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

XLSX = r"D:\USER\DESKTOPP\excel\MAPA GERAL CPI-7 - 2026.xlsx"
wb = openpyxl.load_workbook(XLSX, data_only=True)

UNITS = ["CPI7", "7BPMI", "12BPMI", "22BPMI", "40BPMI", "50BPMI", "53BPMI", "54BPMI", "55BPMI", "14BAEP"]

print(f"{'ABA':<10} {'LINHAS':<8} {'COM PREFIXO':<15} {'MT':<5} {'CR':<5} {'COM DATA':<10} {'SEM DATA':<10}")
for name in UNITS:
    ws = wb[name]
    total_rows = 0
    com_prefixo = 0
    mt = cr = 0
    com_data = sem_data = 0
    ult_linha = 0
    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row=row, column=1).value
        if seq is None: continue
        total_rows += 1
        tipo = ws.cell(row=row, column=2).value
        prefixo = ws.cell(row=row, column=5).value
        data_baixa = ws.cell(row=row, column=4).value
        if prefixo:
            com_prefixo += 1
            ult_linha = row
            if tipo == 'MT': mt += 1
            elif tipo == 'CR': cr += 1
            if data_baixa: com_data += 1
            else: sem_data += 1
    print(f"{name:<10} {total_rows:<8} {com_prefixo:<15} {mt:<5} {cr:<5} {com_data:<10} {sem_data:<10} (ult={ult_linha})")

print()
print("Total estimado de viaturas:", sum(
    sum(1 for row in range(2, wb[n].max_row + 1)
        if wb[n].cell(row=row, column=5).value)
    for n in UNITS
))
